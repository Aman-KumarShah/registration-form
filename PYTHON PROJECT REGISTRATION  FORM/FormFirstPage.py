from tkinter import *
from PIL import Image,ImageTk
from tkinter import ttk
import tkinter.messagebox as tmsg
import openpyxl
from datetime import datetime

root=Tk()
root.title("BrainHacks")
root.wm_iconbitmap("stceticon.ico")

head_menu=Menu(root)
root.configure(menu=head_menu)

Width = root.winfo_screenwidth()
Height = root.winfo_screenheight()
root.geometry(f"{Width}x{Height}+-6+0")
root.minsize(Width, Height)

flag=True
count=0 
flag1=True
count1=0

workbook = openpyxl.load_workbook('FormData.xlsx')
worksheet = workbook.active
Max_row = worksheet.max_row

def function(e):
    workbook = openpyxl.load_workbook('FormData.xlsx')
    worksheet = workbook.active
    if worksheet.max_row > Max_row:
        b2.configure(state=NORMAL)
        b2.bind('<Button-1>',third_page)

root.bind('<Enter>',function)

def second_page(event):
    import FormSecondPage
    FormSecondPage.second_page(root)

def third_page(event):
    import FormThirdPage
    FormThirdPage.third_page(root)

def next_image():
    global count
    if flag==True:
        count+=1
        if count>len(image_list)-1:
            count=0
        l0.configure(image=image_list[count])
        l0.after(4000,next_image)
def next_text():
    global count1
    if flag1==True:
        count1+=1
        if count1>len(text_list)-1:
            count1=0
        l1.configure(text=text_list[count1])
        l1.after(4000,next_text)
def freeze(event):
    global flag
    flag=False
def start(event):
    global flag
    flag=True
    l0.after(4000,next_image)
def freeze1(event):
    global flag1
    flag1=False
def start1(event):
    global flag1
    flag1=True
    l1.after(4000,next_text)

def give_rate():
    rate_root=Toplevel()
    rate_root.geometry("630x520+5+60")
    rate_root.resizable(0,0)
    rate_root.iconbitmap("stceticon.ico")
    rate_root.title("Feedback")
    rate_root.configure(borderwidth=5,relief=SUNKEN,bg="alice blue")
    l=Label(rate_root,text="HELP US IMPROVE",font="widelatin 40 bold underline",fg="#17202A",bg="alice blue")
    l.place(x=5,y=5,width=600,height=70)
    l1=Label(rate_root,text="Please tell us what you think",font="widelatin 20 bold",fg="#17202A",bg="alice blue")
    l1.place(x=0,y=85,width=600,height=30)
    l2=Label(rate_root,text="Name:",font="lucida 15 bold",fg="#17202A",bg="alice blue")
    l2.place(x=35,y=130)

    l3=Label(rate_root,text="Email:",font="lucida 15 bold",fg="#17202A",bg="alice blue")
    l3.place(x=340,y=130)

    def submit():
        if(e1.get()=="" or e2.get()=="" or t.get(1.0, "end-1c")==""):
            tmsg.showerror("Error","Please fill up all the fields")
        else:
            workbook=openpyxl.load_workbook("FeedbackData.xlsx")
            ws=workbook.active
            now = datetime.now()
            dat = now.strftime("%d/%m/%Y %H:%M:%S")
            ws.append([e1.get(),e2.get(),t.get(1.0, "end-1c"),dat])
            workbook.save("FeedbackData.xlsx")
            tmsg.showinfo("Submitted","Successfully Submitted\nThank you for your feedback.")
            e1.delete(0,END)
            e2.delete(0,END)
            t.delete(1.0,END)



    def clear():
        l=tmsg.askquestion("clear","Do you want to clear")
        if(l=="yes"):
            e1.delete(0,END)
            e2.delete(0,END)
            t.delete(1.0,END)

    def on_closing():
        if(e1.get()!="" or e2.get()!="" or t.get(1.0, "end-1c")!=""):
            a=tmsg.askyesno("Warning","Do you want to exit.\nYour feedback is not yet submitted.")
            if(a==True):
                rate_root.destroy()
        else:
            rate_root.destroy()

    e1 = Entry(rate_root, text='', font="times 18",relief=FLAT,fg="black", bg="white", highlightbackground="grey", highlightcolor="#0a75bf", highlightthickness=1)
    # e1=ttk.Entry(rate_root,font="times 20")
    e1.place(x=38,y=165,width=220,height=35)

    # e2=ttk.Entry(rate_root,font="times 20 ")
    e2 = Entry(rate_root, text='', font="times 18",relief=FLAT,fg="black", bg="white", highlightbackground="grey", highlightcolor="#0a75bf", highlightthickness=1)
    e2.place(x=340,y=165,width=240,height=35)

    l4=Label(rate_root,text="Comment:",font="lucida 15 bold",fg="#17202A",bg="alice blue")
    l4.place(x=30,y=220)
    t=Text(rate_root,font="times 18")
    t.place(x=30,y=260,width=560,height=210)
    scroll=Scrollbar(t,command=t.yview)
    scroll.place(relheight=1,relx=0.974)
    t.configure(yscrollcommand=scroll.set)
    b1=ttk.Button(rate_root,text="SUMBIT",comman=submit,cursor="hand2")
    b1.place(x=100,y=480)
    b1=ttk.Button(rate_root,text="CLEAR",command=clear,cursor="x_cursor")
    b1.place(x=420,y=480)
    rate_root.grab_set()
    rate_root.protocol("WM_DELETE_WINDOW", on_closing)
    # printstring()
    rate_root.mainloop()

def support():
    support_root = Toplevel()
    support_root.title("Chatbot")
    support_root.geometry("495x513+60+60")
    support_root.resizable(0,0)
    support_root.iconbitmap("stceticon.ico")
    BG_GRAY = "#ABB2B9"
    BG_COLOR = "#17202A"
    TEXT_COLOR = "#EAECEE"
    support_root.configure(borderwidth=5,relief=SUNKEN,bg=BG_COLOR)
 
    FONT = "Helvetica 17"
    FONT_BOLD = "Helvetica 15 bold"
    def send(event=support_root):
        txt.configure(state=NORMAL)
        send = "You : " + e.get()
        txt.insert(END, send+"\n\n")
 
        user = e.get().lower()
 
        if (user == "hello"):
            txt.insert(END, "Bot : Hi there, how can I help?")
 
        elif (user == "hi" or user == "hii" or user == "hiiii"):
            txt.insert(END, "Bot : Hi there, what can I do for you?")
 
        elif (user.find("how are you")!=-1):
            txt.insert(END, "Bot : fine! and you")
 
        elif (user == "fine" or user == "i am good" or user == "i am doing good"):
            txt.insert(END, "Bot : Great! how can I help you.")
 
        elif (user.find("thanks")!=-1 or user.find("thank you")!=-1):
            txt.insert(END, "Bot : My pleasure !")
 
        elif (user == "tell me a joke" or user == "tell me something funny" or user == "crack a funny line"):
            txt.insert(END, "Bot : What did the buffalo say\nwhen his son left for college? Bison.! ")
    
        elif (user.find("what is brainhacks")!=-1 or user.find("what is this competition")!=-1 or user.find("what is the prize")!=-1 or user.find("what is the criteria of this competition")!=-1 or user.find("what is the website")!=-1 or user.find("website")!=-1 or user.find("how to fill up the form")!=-1 or user.find("it it an online competition")!=-1 or user.find("it it an offline competition")!=-1):
            txt.insert(END, "Bot : Please vist our website\n'www.brainhackscoding.org.in'")
        
        elif(user.find("contact")!=-1 or user.find("helpline")!=-1):
            txt.insert(END, "Bot : 'www.brainhackscoding.org.in'")
            
        elif(user.find("what")!=-1 or user.find("how")!=-1):
            txt.insert(END, "Bot : For all such queries do vist\nour website 'www.brainhackscoding.org.in'")
        elif (user.find("help")!=-1):
            txt.insert(END, "Bot : Yeah,I will surely guide you.!")

        elif (user.find("i have some problem with the form fill")!=-1 or user.find("there is a problem with my registration")!=-1 or user.find("i have a query")!=-1 or user.find("there is a query")!=-1 or user.find("there is some issues")!=-1 or user.find("i have a problem")!=-1 or user.find("there is an issue")!=-1 or user.find("i have some issue")!=-1 or user.find("i have some problem")!=-1 or user.find("i have issue ")!=-1 or user.find("i have an issue ")!=-1):
            txt.insert(END, "Bot : We are very sorry for the inconvenience.\n")
            txt.insert(END, "Bot : Do mail your query(s) to us at\nbrainhacks.code@gmail.com")
    
        elif (user.find("by when my problem will be solved")!=-1 or user.find("by when will my")!=-1 or user.find("when my issue will be resolved ")!=-1 or user.find("please look after my issue fast")!=-1 or user.find("i have mailed my issue")!=-1 or user.find("i have mailed my query")!=-1 or user.find("i have mailed my question")!=-1 or user.find("i have posted my question")!=-1 or user.find("solve it fast")!=-1 or user.find("by what time my issue")!=-1):
            txt.insert(END, "Bot : We are on it\n")
            txt.insert(END, "Bot : Please be assured that your issue\nwill be resolved within 2-3 working days.")
 
        elif (user.find("goodbye")!=-1 or user.find("bye")!=-1 or user == " ok see yaa" or user == "see yaa"):
            txt.insert(END, "Bot : Good bye Have a nice day!")
            e.delete(0,END)
        
        elif(user=="ok"):
            txt.insert(END, "Bot : Yes")
        else:
            txt.insert(END, "Bot : Sorry! I didn't understand that")
        txt.insert(END,"\n\n")
        txt.configure(state=DISABLED)
        if(user.find("goodbye")!=-1 or user.find("bye")!=-1 or user == " ok see yaa" or user == "see yaa"):
            e.delete(0,END)
            support_root.destroy()
        else:
            e.delete(0, END)
            txt.see(END)

    def enter(event):
        user = e.get()
        if(user=="Say Hello..." or user=="Write here Something..."):
            e.delete(0, END)

    def leave(event):
        user = e.get().lower()
        if(user=="" and txt.get(1.0,"end-1c")==""):
            e.insert(END,"Say Hello...")
        elif(user=="" and txt.get(1.0,"end-1c")!="" ):
            e.insert(END,"Write here Something...")
 
    lable1 = Label(support_root, bg=BG_COLOR, fg=TEXT_COLOR, text="Welcome", font="times 25 bold").place(x=185,y=4)
 
    txt = Text(support_root, bg="#11364f", fg=TEXT_COLOR, font=FONT, width=37,height=16,state=DISABLED)
    txt.place(x=0,y=45)
    
    scrollbar = Scrollbar(txt)
    scrollbar.place(relheight=1, relx=0.974)
    txt.configure(yscrollcommand=scrollbar.set)    
    scrollbar.configure(command=txt.yview)
 
    e = Entry(support_root, bg="#2C3E50", fg=TEXT_COLOR, font="Helvetica 18",width=32)
    e.place(x=0,y=470)

    e.insert(END,"Say Hello...")
    e.bind("<Button-1>",enter)
    e.bind("<FocusOut>",leave)
    e.bind("<Return>",send)
 
    send = Button(support_root, text="Send", font="Helvetica 14 bold", bg=BG_GRAY,
              command=send).place(x=423,y=467)

    # printstring()
    support_root.mainloop()



head_menu.add_command(label="Feedback",command=give_rate)
head_menu.add_command(label="Support",command=support)



f0=Frame(root,width=Width//2,borderwidth=2,relief=SUNKEN,bg='white')
f0.pack(fill=Y,side=LEFT)

heading_frame = Frame(root,width=Width//2-10,height=Height//2-120,borderwidth=2,relief=GROOVE,bg='mint cream')
heading_frame.pack()

l1 = Label(heading_frame,text="BRAINHACKS",font="lucida 50 bold",bg="mint cream")
l1.place(x=(Width//2-10)//5,y=30)

l3=Label(heading_frame,text="It all begins here...",font="lucida 18 bold",bg="mint cream")
l3.place(x=(Width//2-10)//5+230,y=100)

style = ttk.Style()
style.configure('TButton',background='mint cream')
b1=ttk.Button(heading_frame,text="Application Form Fill-Up  ",width=25,cursor="hand2")
b1.bind("<Button-1>",second_page)
b1.place(x=(Width//2)-390,y=heading_frame['height']-50)

b2=ttk.Button(heading_frame,text="View Your Filled Application Form",width=30,cursor="hand2",state=DISABLED)
b2.place(x=(Width//2)-220,y=heading_frame['height']-50)

# def exit(event):
#     if tmsg.askyesno('Exit','Do you want to leave the page?'):
#         root.destroy()

def close_window():
    if tmsg.askyesno('Exit','Do you want to leave the page?'):
        root.destroy()

b3=ttk.Button(heading_frame,text="Exit",width=8,cursor="hand2")
b3.place(x=15,y=heading_frame['height']-50)
b3.bind('<Button-1>',quit)
root.protocol("WM_DELETE_WINDOW",close_window)

f1=Frame(root,width=Width//2,height=Height//2+40,borderwidth=2,relief=SUNKEN,bg='white')
f1.pack()

image_list=[]
for i in range(1,7):
    image=Image.open(f"imageA{i}.png").resize((Width//2,Height))
    photo=ImageTk.PhotoImage(image)
    image_list.append(photo)
    
l0=ttk.Label(f0,image=image_list[0])
l0.pack(fill=Y)
l0.after(4000,next_image)
l0.bind("<Enter>",freeze)
l0.bind("<Leave>",start)


text_list=[]
for i in range(1,5):
    f=open(f"textA{i}.txt","r")
    str=f.read()
    text_list.append(str)
imt=Image.open("imagefront.png").resize((Width//2-15,Height//2+22))
pht=ImageTk.PhotoImage(imt)
lmt=Label(f1,image=pht,borderwidth=2,relief=SUNKEN)
lmt.place(x=0,y=0)

l1=Label(f1,text=text_list[0],bg="#000027",font="cambria 12 bold",fg="white",borderwidth=3,relief=SUNKEN,justify=LEFT)
l1.place(x=10,y=10)
l1.after(4000,next_text)
l1.bind("<Enter>",freeze1)
l1.bind("<Leave>",start1)


root.mainloop()